import streamlit as st
import pandas as pd
import numpy as np
import io
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import xlsxwriter

# ==============================================================================
# PAGE CONFIGURATION & UI CLEANUP
# ==============================================================================
st.set_page_config(page_title="Automated Reporting Portal", page_icon="📊", layout="centered")

hide_streamlit_branding = """
    <style>
    #MainMenu {visibility: hidden;}
    header {visibility: hidden;}
    footer {visibility: hidden;}
    a[href*="github.com"] {display: none !important;}
    </style>
"""
st.markdown(hide_streamlit_branding, unsafe_allow_html=True)


# ==============================================================================
# SECURITY: Basic Password Authentication
# ==============================================================================
def check_password():
    def password_entered():
        if st.session_state["password"] == "Operaciones2026!":
            st.session_state["password_correct"] = True
            del st.session_state["password"] 
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        st.text_input("🔒 Please enter the password to access the portal:", type="password", on_change=password_entered, key="password")
        return False
    elif not st.session_state["password_correct"]:
        st.text_input("❌ Incorrect password. Please try again:", type="password", on_change=password_entered, key="password")
        return False
    return True


# ==============================================================================
# HELPER FUNCTIONS 
# ==============================================================================
def get_rounded_time_display():
    """
    Calculates time display in CDMX Timezone (UTC-6) according to rounding rules:
    IF MINUTE <= 15 -> :00
    IF MINUTE <= 45 -> :30
    ELSE -> next hour :00
    """
    cdmx_tz = datetime.timezone(datetime.timedelta(hours=-6))
    now = datetime.datetime.now(cdmx_tz)
    minute = now.minute
    if minute <= 15:
        r_dt = now.replace(minute=0, second=0, microsecond=0)
    elif minute <= 45:
        r_dt = now.replace(minute=30, second=0, microsecond=0)
    else:
        r_dt = (now + datetime.timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
    
    return r_dt.strftime("%I:%M %p").lstrip('0')

def clean_spv_value(val):
    """
    Consolidates CEDIS, NaN, empty strings, and unmapped rows into #N/A
    """
    val_str = str(val).strip() if pd.notna(val) else '#N/A'
    if val_str in ['CEDIS', 'nan', 'NaN', '', 'None', 'NoneType']:
        return '#N/A'
    return val_str

def clean_spv_df(spv_file_bytes):
    """
    Standardizes the SUPERVISORES mapping file to ensure SPV, ZONA, and PDV columns 
    are universally correctly cased and formatted to prevent KeyErrors.
    """
    df_spv = pd.read_excel(io.BytesIO(spv_file_bytes))
    df_spv.columns = [str(c).strip().upper() for c in df_spv.columns]
    
    if 'SUPERVISOR' in df_spv.columns and 'SPV' not in df_spv.columns:
        df_spv.rename(columns={'SUPERVISOR': 'SPV'}, inplace=True)
        
    if 'ZONA' not in df_spv.columns:
        df_spv['ZONA'] = ''
        
    if 'PDV' not in df_spv.columns:
        df_spv['PDV'] = ''
        
    return df_spv


# ==============================================================================
# REPORT LOGIC WRAPPERS
# ==============================================================================

def generate_tiktok_visits(raw_file_bytes, spv_file_bytes):
    # 1. SAFELY LOAD DATA INTO MEMORY
    xls_raw = pd.ExcelFile(io.BytesIO(raw_file_bytes))
    original_sheet_name = '订单综合信息' if '订单综合信息' in xls_raw.sheet_names else xls_raw.sheet_names[0]
    df_raw = pd.read_excel(xls_raw, sheet_name=original_sheet_name)
    
    # Clean redundant calculated columns from previous runs if present in raw data sheet
    for col in ['SPV', 'SLR']:
        if col in df_raw.columns:
            df_raw = df_raw.drop(columns=[col])
            
    df_raw_copy = df_raw.copy()  # Preserve original raw data to write back
    
    # Load supervisor mapping file
    df_sup = clean_spv_df(spv_file_bytes)
    df_sup.rename(columns={'SPV': 'SUPERVISOR'}, inplace=True)
            
    if 'PDV' in df_sup.columns:
        df_sup['PDV'] = df_sup['PDV'].astype(str).str.strip()
    
    if 'Punto de Recogida' in df_raw.columns:
        df_raw['Punto de Recogida'] = df_raw['Punto de Recogida'].astype(str).str.strip()

    # 2. AUTOMATIC TARGET HOUR DETECTION FROM 'TABLA' SHEET
    try:
        df_tabla = pd.read_excel(xls_raw, sheet_name='TABLA')
        existing_cols = [str(c).strip() for c in df_tabla.columns]
        
        if '9' not in existing_cols: target_hour = 9
        elif '12' not in existing_cols: target_hour = 12
        elif '3' not in existing_cols: target_hour = 3
        else: target_hour = 5
    except Exception:
        df_tabla = pd.DataFrame()
        target_hour = 9

    # 3. DYNAMIC DATE EXTRACTION & +1 DAY SHIFT
    months_en = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun', 
                 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
    
    report_date_str = ""
    if 'Tiempo de registro de la orden' in df_raw.columns:
        dates = pd.to_datetime(df_raw['Tiempo de registro de la orden']).dt.date
        o_date = dates.value_counts().idxmax()
        
        # Add 1 day to date for pickup day representation
        target_reco_date = o_date + datetime.timedelta(days=1)
        report_date_str = f"{target_reco_date.day:02d}-{months_en[target_reco_date.month]}"
        
        # Filter dataset strictly to orders from the order date
        df_raw = df_raw[dates == o_date].copy()
    else:
        report_date_str = "N/A"

    # 4. DATA CLEANING & MERGING
    df_raw['SLR'] = df_raw['Compañía remitente'].astype(str).apply(
        lambda x: str(x).split('|')[-1].strip() if '|' in str(x) else str(x).strip()
    )
    
    df_raw = pd.merge(df_raw, df_sup[['PDV', 'SUPERVISOR']], left_on='Punto de Recogida', right_on='PDV', how='left')
    df_raw.rename(columns={'SUPERVISOR': 'SPV'}, inplace=True)
    
    # Explicitly consolidate CEDIS / NaNs / Empty into '#N/A'
    df_raw['SPV'] = df_raw['SPV'].apply(clean_spv_value)
    
    # Filter out both "CANCELADO" and "EN RECEPCIÓN"
    if 'Estado de la orden' in df_raw.columns:
        df_raw['Estado_upper'] = df_raw['Estado de la orden'].astype(str).str.upper().str.strip()
        df_valid = df_raw[~df_raw['Estado_upper'].isin(['CANCELADO', 'EN RECEPCIÓN', 'EN RECEPCION'])]
    else:
        df_valid = df_raw.copy()
        
    # Remove duplicate tracking numbers to prevent inflation of package counts
    if 'Guía de Rastreo' in df_valid.columns:
        df_valid = df_valid.drop_duplicates(subset=['Guía de Rastreo']).copy()

    # 5. TIME LOGIC & PIVOTS
    time_col = str(target_hour)
    df_current_counts = df_valid.groupby('SLR')['Guía de Rastreo'].count().reset_index(name=time_col)

    if target_hour == 9:
        df_mapping = df_valid[['SLR', 'Punto de Recogida', 'SPV']].drop_duplicates(subset=['SLR'])
        df_mapping.rename(columns={'Punto de Recogida': 'PDV'}, inplace=True)
        
        df_tabla = pd.merge(df_mapping, df_current_counts, on='SLR', how='left')
        df_tabla[time_col] = df_tabla[time_col].fillna(0).astype(int)
        df_tabla['SPV'] = df_tabla['SPV'].apply(clean_spv_value)
        
        df_template_source = df_tabla.copy()
        vol_col = time_col
    else:
        df_tabla.columns = [str(c).strip() for c in df_tabla.columns]
        if time_col in df_tabla.columns: df_tabla.drop(columns=[time_col], inplace=True)
        if 'FORMULA' in df_tabla.columns: df_tabla.drop(columns=['FORMULA'], inplace=True)

        df_tabla = pd.merge(df_tabla, df_current_counts, on='SLR', how='left')
        df_tabla['SPV'] = df_tabla['SPV'].apply(clean_spv_value)
        
        # 0s in time slots 12, 3, 5 mean all packages collected -> converted to '#N/A'
        df_tabla[time_col] = df_tabla[time_col].apply(lambda x: '#N/A' if pd.isna(x) or x == 0 else int(x))
        
        if '9' in df_tabla.columns:
            def calculate_formula(row):
                val_9 = row['9']
                val_curr = row[time_col]
                if str(val_curr) == '#N/A' or str(val_9) == '#N/A':
                    return '#N/A'
                try:
                    diff = int(val_9) - int(val_curr)
                    return diff
                except Exception:
                    return '#N/A'

            df_tabla['FORMULA'] = df_tabla.apply(calculate_formula, axis=1)
        else:
            raise ValueError("Error: Columna '9' no encontrada en la hoja TABLA. Procesa primero las 9 AM.")
            
        # Filter for sellers where formula results in 0
        df_template_source = df_tabla[df_tabla['FORMULA'].astype(str) == '0'].copy()
        vol_col = time_col

    # REORDER TABLA COLUMNS: SPV -> PDV -> SLR -> hours (9, 12, 3, 5, FORMULA, etc.)
    desired_order = ['SPV', 'PDV', 'SLR']
    other_columns = [col for col in df_tabla.columns if col not in desired_order]
    final_tabla_columns = [col for col in desired_order if col in df_tabla.columns] + other_columns
    df_tabla = df_tabla[final_tabla_columns]

    # 6. TEMPLATE GENERATION WITH CUSTOM SORTING (#N/A AT THE VERY END)
    unique_spvs = [s for s in df_template_source['SPV'].unique() if str(s).strip() != '#N/A']
    unique_spvs.sort()
    if '#N/A' in df_template_source['SPV'].unique() or '#N/A' in df_tabla['SPV'].unique():
        unique_spvs.append('#N/A')

    template_data = []
    for spv in unique_spvs:
        group = df_template_source[df_template_source['SPV'] == spv]
        group_vols = pd.to_numeric(group[vol_col], errors='coerce').fillna(0)
        
        template_data.append({
            'SUPERVISOR': spv,
            'TOTAL TT SELLERS': len(group),
            '1-10': len(group[(group_vols >= 1) & (group_vols <= 10)]),
            '11-50': len(group[(group_vols >= 11) & (group_vols <= 50)]),
            '+50': len(group[group_vols > 50])
        })
        
    df_template = pd.DataFrame(template_data)
    
    total_row = {
        'SUPERVISOR': 'TOTAL',
        'TOTAL TT SELLERS': df_template['TOTAL TT SELLERS'].sum() if not df_template.empty else 0,
        '1-10': df_template['1-10'].sum() if not df_template.empty else 0,
        '11-50': df_template['11-50'].sum() if not df_template.empty else 0,
        '+50': df_template['+50'].sum() if not df_template.empty else 0
    }
    df_template = pd.concat([df_template, pd.DataFrame([total_row])], ignore_index=True)

    # 7. EXCEL OUTPUT GENERATION
    output_buffer = io.BytesIO()

    with pd.ExcelWriter(output_buffer, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}}) as writer:
        workbook = writer.book
        
        # --- Sheet 1: VISITAS ---
        sheet_name = 'VISITAS'
        worksheet = workbook.add_worksheet(sheet_name)
        worksheet.hide_gridlines(0)
        
        font_family = 'Times New Roman'
        
        blue_bg = workbook.add_format({'bg_color': '#5B9BD5', 'font_color': 'black', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_name': font_family, 'font_size': 13})
        purple_bg = workbook.add_format({'bg_color': '#7030A0', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'text_wrap': True, 'font_name': font_family, 'font_size': 13})
        
        # DATE FORMAT SET TO 16PT BOLD
        purple_date_bg = workbook.add_format({'bg_color': '#7030A0', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_name': font_family, 'font_size': 16})
        
        red_bg = workbook.add_format({'bg_color': '#C00000', 'font_color': 'white', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_name': font_family, 'font_size': 13})
        yellow_bg = workbook.add_format({'bg_color': '#FFFF00', 'font_color': 'black', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'bold': True, 'font_name': font_family, 'font_size': 13})
        
        title_fmt = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter', 'font_name': font_family})
        sub_title_fmt = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center', 'valign': 'vcenter', 'font_name': font_family})
        time_badge_fmt = workbook.add_format({'bg_color': '#FFFF00', 'bold': True, 'italic': True, 'align': 'center', 'valign': 'vcenter', 'font_size': 18, 'font_name': font_family, 'border': 1})
        
        data_fmt = workbook.add_format({'border': 1, 'align': 'left', 'valign': 'vcenter', 'font_name': font_family, 'font_size': 11})
        data_num_fmt = workbook.add_format({'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_name': font_family, 'font_size': 11})
        
        worksheet.set_column('A:A', 54)
        worksheet.set_column('B:E', 24)

        worksheet.merge_range('B1:D1', 'TT SELLERS TO VISIT BASED ON THE # OF PACKAGES', title_fmt)
        worksheet.merge_range('B2:D2', '按包裹数量需拜访的TT卖家', sub_title_fmt)
        
        formatted_time = get_rounded_time_display()
        worksheet.merge_range('E1:E2', formatted_time, time_badge_fmt)
        worksheet.set_row(0, 26)
        worksheet.set_row(1, 26)
        
        worksheet.write(2, 0, 'SUPERVISOR', blue_bg)
        worksheet.write(2, 1, 'TOTAL TT SELLERS', purple_bg)
        worksheet.write(2, 2, '1-10', red_bg)
        worksheet.write(2, 3, '11-50', red_bg)
        worksheet.write(2, 4, '+50', red_bg)
        worksheet.set_row(2, 32)
        
        # BOLD 16PT DATE IN ROW 4
        worksheet.write(3, 0, '', purple_date_bg)
        worksheet.merge_range(3, 1, 3, 4, report_date_str, purple_date_bg)
        worksheet.set_row(3, 26)
        
        current_row = 4
        for _, row in df_template.iterrows():
            is_total_row = (row['SUPERVISOR'] == 'TOTAL')
            row_format = yellow_bg if is_total_row else data_num_fmt
            str_format = yellow_bg if is_total_row else data_fmt
            
            # Blank zero values on individual data rows (except Total)
            v_tot = row['TOTAL TT SELLERS']
            v_1_10 = row['1-10']
            v_11_50 = row['11-50']
            v_50 = row['+50']
            
            disp_tot = int(v_tot) if is_total_row or v_tot != 0 else ""
            disp_1_10 = int(v_1_10) if is_total_row or v_1_10 != 0 else ""
            disp_11_50 = int(v_11_50) if is_total_row or v_11_50 != 0 else ""
            disp_50 = int(v_50) if is_total_row or v_50 != 0 else ""

            worksheet.write(current_row, 0, str(row['SUPERVISOR']), str_format)
            worksheet.write(current_row, 1, disp_tot, row_format)
            worksheet.write(current_row, 2, disp_1_10, row_format)
            worksheet.write(current_row, 3, disp_11_50, row_format)
            worksheet.write(current_row, 4, disp_50, row_format)
            worksheet.set_row(current_row, 22)
            current_row += 1

        # --- Sheet 2: TABLA ---
        df_tabla.to_excel(writer, sheet_name='TABLA', index=False)
        worksheet_tabla = writer.sheets['TABLA']
        worksheet_tabla.hide_gridlines(0)
        
        tabla_header = workbook.add_format({'bold': True, 'bg_color': '#D9E1F2', 'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_name': font_family, 'font_size': 12})
        tabla_data = workbook.add_format({'border': 1, 'valign': 'vcenter', 'font_name': font_family, 'font_size': 11})
        
        for col_num, value in enumerate(df_tabla.columns):
            worksheet_tabla.write(0, col_num, value, tabla_header)
            val_lens = df_tabla[value].astype(str).str.len()
            max_val_len = val_lens.max() if not val_lens.empty and pd.notna(val_lens.max()) else 0
            max_len = max(max_val_len, len(str(value))) + 4
            worksheet_tabla.set_column(col_num, col_num, max(max_len, 14), tabla_data)
        worksheet_tabla.set_row(0, 26)

        # --- Sheet 3: ORIGINAL RAW DATA (AUTO-STRETCHED COLUMNS) ---
        df_raw_copy.to_excel(writer, sheet_name=original_sheet_name, index=False)
        worksheet_raw = writer.sheets[original_sheet_name]
        worksheet_raw.hide_gridlines(0)
        
        raw_header_fmt = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#E0E0E0', 'font_name': font_family, 'font_size': 11})
        raw_data_fmt = workbook.add_format({'border': 1, 'font_name': font_family, 'font_size': 11})
        
        for col_num, col_col in enumerate(df_raw_copy.columns):
            worksheet_raw.write(0, col_num, col_col, raw_header_fmt)
            val_lens = df_raw_copy[col_col].astype(str).str.len()
            max_val_len = val_lens.max() if not val_lens.empty and pd.notna(val_lens.max()) else 0
            max_len = max(max_val_len, len(str(col_col))) + 4
            worksheet_raw.set_column(col_num, col_num, max(max_len, 12), raw_data_fmt)
        worksheet_raw.set_row(0, 24)

    output_filename = "TIKTOK PENDING FROM YESTERDAY.xlsx"
    return output_buffer.getvalue(), output_filename


def generate_aliexpress(raw_file_bytes, spv_file_bytes):
    df_spv = clean_spv_df(spv_file_bytes)
    df_spv['PDV'] = df_spv['PDV'].astype(str).str.strip()
    
    valid_spvs = [str(spv).strip().upper() for spv in df_spv['SPV'].dropna().unique() if 'CEDIS' not in str(spv).strip().upper()]

    try:
        xls = pd.ExcelFile(io.BytesIO(raw_file_bytes))
        original_sheet_name = '订单综合信息' if '订单综合信息' in xls.sheet_names else xls.sheet_names[0]
        df_raw = pd.read_excel(xls, sheet_name=original_sheet_name)
    except ValueError:
        df_raw = pd.read_excel(io.BytesIO(raw_file_bytes))
        original_sheet_name = 'Sheet1'

    df_raw_original = df_raw.copy()

    if 'Estado de la orden' in df_raw.columns:
        df_raw = df_raw[df_raw['Estado de la orden'] != 'Cancelado']

    seller_pdv_map = df_raw[['Compañía remitente', 'Punto de Recogida']].dropna().drop_duplicates(subset=['Compañía remitente'])
    seller_pdv_map = seller_pdv_map.rename(columns={'Punto de Recogida': 'PDV'})
    seller_pdv_map['PDV'] = seller_pdv_map['PDV'].astype(str).str.strip()

    df_total = df_raw.groupby('Compañía remitente')['Guía de Rastreo'].count().reset_index(name='Total')
    df_abnormal = df_raw[df_raw['Estado de la orden'] == 'La recogida falló'].groupby('Compañía remitente')['Guía de Rastreo'].count().reset_index(name='Abnormal')

    pending_statuses = ['Asignado a PDV', 'Asignado a un mensajero', 'Asignado a mensajero', 'Pendiente', 'Creado']
    df_por_rec = df_raw[df_raw['Estado de la orden'].isin(pending_statuses)].groupby('Compañía remitente')['Guía de Rastreo'].count().reset_index(name='PorRec')

    df_report = df_total.copy()
    df_report = pd.merge(df_report, df_abnormal, on='Compañía remitente', how='left')
    df_report = pd.merge(df_report, df_por_rec, on='Compañía remitente', how='left')

    df_report['Total'] = df_report['Total'].fillna(0).astype(int)
    df_report['Abnormal'] = df_report['Abnormal'].fillna(0).astype(int)
    df_report['PorRec'] = df_report['PorRec'].fillna(0).astype(int)
    df_report['Visita'] = df_report['Total'] - df_report['Abnormal'] - df_report['PorRec']

    df_report = pd.merge(df_report, seller_pdv_map, on='Compañía remitente', how='left')
    df_report = pd.merge(df_report, df_spv[['PDV', 'SPV', 'ZONA']], on='PDV', how='left')

    df_report = df_report.rename(columns={'Compañía remitente': 'SELLER'})
    df_report = df_report.dropna(subset=['SPV']) 
    df_report = df_report[df_report['SPV'].str.strip() != '']

    df_report['SPV'] = df_report['SPV'].astype(str).str.upper()
    df_report['ZONA'] = df_report['ZONA'].fillna('').astype(str)
    df_report['PDV'] = df_report['PDV'].fillna('').astype(str)
    df_report['SELLER'] = df_report['SELLER'].fillna('').astype(str)
    
    df_report = df_report[df_report['SPV'].isin(valid_spvs)]
    df_report = df_report.sort_values(by=['SPV', 'ZONA', 'PDV', 'SELLER'], ascending=[True, True, True, True])

    df_report['Rate'] = np.where(df_report['Total'] > 0, (df_report['Visita'] + df_report['Abnormal']) / df_report['Total'], 0.0)
    df_report['Abnormal_View'] = df_report['Abnormal'].apply(lambda x: '' if x == 0 else int(x))
    df_report['PorRec_View'] = df_report['PorRec'].apply(lambda x: '' if x == 0 else int(x))

    months_es = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}

    if 'Tiempo de registro de la orden' in df_raw.columns:
        dates = pd.to_datetime(df_raw['Tiempo de registro de la orden']).dt.date
        o_date = dates.value_counts().idxmax() 
        r_date = o_date + datetime.timedelta(days=1)
        
        if r_date.weekday() == 0: 
            o_date_prev = o_date - datetime.timedelta(days=1)
            o_date_zh = f"{o_date_prev.month}月{o_date_prev.day}日 & {o_date.month}月{o_date.day}日"
            o_date_es = f"{months_es[o_date_prev.month]} {o_date_prev.day} & {months_es[o_date.month]} {o_date.day}"
        else:
            o_date_zh = f"{o_date.month}月{o_date.day}日"
            o_date_es = f"{months_es[o_date.month]} {o_date.day}"
            
        r_date_zh = f"{r_date.month}月{r_date.day}日"
        r_date_es = f"{months_es[r_date.month]} {r_date.day}"
    else:
        o_date_zh, o_date_es = "7月29日", "Jul 29"
        r_date_zh, r_date_es = "7月30日", "Jul 30"

    subtitle_str = f"订单日期：{o_date_zh}, 揽收日期 {r_date_zh} Pedidos: {o_date_es} | Reco: {r_date_es}"

    mx_tz = datetime.timezone(datetime.timedelta(hours=-6))
    now = datetime.datetime.now(mx_tz)
    ts = now.timestamp()
    ts_rounded = round(ts / 900) * 900
    now_rounded = datetime.datetime.fromtimestamp(ts_rounded, mx_tz)
    time_str = f"{now_rounded.hour}" if now_rounded.minute == 0 else f"{now_rounded.hour}.{now_rounded.minute:02d}"

    output_filename = f"ALI {r_date_es.upper()} AT {time_str}.xlsx"

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer.book
    worksheet = workbook.add_worksheet('AliExpress')

    magenta, yellow, red = '#B0005B', '#FFFF00', '#FF0000'
    base_font = 'Calibri'
    base_size = 14

    title_format = workbook.add_format({'bold': True, 'font_size': 36, 'align': 'center', 'valign': 'vcenter', 'bg_color': magenta, 'font_color': 'white', 'font_name': base_font})
    # BOLD 24PT DATE SUBTITLE
    subtitle_format = workbook.add_format({'bold': True, 'font_size': 24, 'align': 'center', 'valign': 'vcenter', 'bg_color': magenta, 'font_color': 'white', 'font_name': base_font})
    header_format = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': magenta, 'font_color': 'white', 'border': 1, 'text_wrap': True, 'font_name': base_font})

    grand_total_label = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': magenta, 'font_color': 'white', 'border': 1, 'font_name': base_font})
    grand_total_num = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': magenta, 'font_color': 'white', 'border': 1, 'num_format': '#,##0', 'font_name': base_font})
    grand_total_pct = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': red, 'font_color': 'white', 'border': 1, 'num_format': '0.00%', 'font_name': base_font})

    subtotal_label_yellow = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': yellow, 'font_color': 'black', 'border': 1, 'font_name': base_font})
    subtotal_num_yellow = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': yellow, 'font_color': 'black', 'border': 1, 'num_format': '#,##0', 'font_name': base_font})
    subtotal_pct_red = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': red, 'font_color': 'white', 'border': 1, 'num_format': '0.00%', 'font_name': base_font})
    subtotal_num_red = workbook.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': red, 'font_color': 'white', 'border': 1, 'num_format': '#,##0', 'font_name': base_font})

    spv_merge_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': True, 'font_size': base_size, 'font_name': base_font})
    data_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': base_size, 'font_name': base_font})
    data_num_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': base_size, 'num_format': '#,##0', 'font_name': base_font})
    data_pct_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': base_size, 'num_format': '0.00%', 'font_name': base_font})

    worksheet.set_column('A:A', 22); worksheet.set_column('B:B', 22); worksheet.set_column('C:C', 36); worksheet.set_column('D:D', 48)
    worksheet.set_column('E:I', 26)

    worksheet.merge_range('A1:I1', 'AliExpress 拜访率 % Visitas AliExpress', title_format)
    worksheet.set_row(0, 80) 
    worksheet.merge_range('A2:I2', subtitle_str, subtitle_format)
    worksheet.set_row(1, 40)

    headers = ['SPV', 'ZONA', '客户归属网点\nPDV', 'SELLER', '待收取总数\nTOTAL a Recolectar', '已记录拜访\nVISITA REGISTRADA', '异常扫描已记录\nABNORMAL SCAN', '待收取包裹数\nPOR RECOLECTAR', '商家拜访率\nRATE %']
    for col, h in enumerate(headers): worksheet.write(2, col, h, header_format)
    worksheet.set_row(2, 60)

    gt_total = int(df_report['Total'].sum())
    gt_abnormal = int(df_report['Abnormal'].sum())
    gt_porrec = int(df_report['PorRec'].sum())
    gt_visita_column = int(df_report['Visita'].sum()) + gt_abnormal
    gt_rate = gt_visita_column / gt_total if gt_total > 0 else 0.0

    worksheet.merge_range('A4:D4', 'GRAND TOTAL 总计', grand_total_label)
    worksheet.write(3, 4, gt_total, grand_total_num)
    worksheet.write(3, 5, gt_visita_column, grand_total_num)
    worksheet.write(3, 6, gt_abnormal, grand_total_num)
    worksheet.write(3, 7, gt_porrec, grand_total_num)
    worksheet.write(3, 8, float(gt_rate), grand_total_pct) 
    worksheet.set_row(3, 34)

    current_row = 4
    for spv_name, group in df_report.groupby('SPV', sort=False):
        spv_start_row = current_row
        
        zona_start_row = current_row
        current_zona = None

        for _, row in group.iterrows():
            if current_zona != row['ZONA']:
                if current_zona is not None:
                    if current_row - 1 > zona_start_row:
                        worksheet.merge_range(zona_start_row, 1, current_row - 1, 1, current_zona, data_format)
                    else:
                        worksheet.write(zona_start_row, 1, current_zona, data_format)
                current_zona = row['ZONA']
                zona_start_row = current_row

            worksheet.write(current_row, 2, row['PDV'], data_format)
            worksheet.write(current_row, 3, row['SELLER'], data_format)
            worksheet.write(current_row, 4, int(row['Total']), data_num_format)
            worksheet.write(current_row, 5, int(row['Visita']), data_num_format)
            worksheet.write(current_row, 6, row['Abnormal_View'], data_num_format)
            worksheet.write(current_row, 7, row['PorRec_View'], data_num_format)
            worksheet.write(current_row, 8, float(row['Rate']), data_pct_format)
            worksheet.set_row(current_row, 24)
            current_row += 1

        if current_zona is not None:
            if current_row - 1 > zona_start_row:
                worksheet.merge_range(zona_start_row, 1, current_row - 1, 1, current_zona, data_format)
            else:
                worksheet.write(zona_start_row, 1, current_zona, data_format)

        if current_row - 1 > spv_start_row: 
            worksheet.merge_range(spv_start_row, 0, current_row - 1, 0, spv_name, spv_merge_format)
        else: 
            worksheet.write(spv_start_row, 0, spv_name, spv_merge_format)

        sub_total, sub_visita, sub_abnormal, sub_porrec = int(group['Total'].sum()), int(group['Visita'].sum()), int(group['Abnormal'].sum()), int(group['PorRec'].sum())
        sub_rate = (sub_visita + sub_abnormal) / sub_total if sub_total > 0 else 0.0

        worksheet.merge_range(f'A{current_row+1}:D{current_row+1}', f'TOTAL {spv_name}', subtotal_label_yellow)
        worksheet.write(current_row, 4, sub_total, subtotal_num_yellow)
        worksheet.write(current_row, 5, sub_visita, subtotal_num_yellow)
        worksheet.write(current_row, 6, sub_abnormal, subtotal_num_yellow)
        worksheet.write(current_row, 7, sub_porrec, subtotal_num_red)
        worksheet.write(current_row, 8, float(sub_rate), subtotal_pct_red)

        worksheet.set_row(current_row, 34)
        current_row += 1

    df_raw_original.to_excel(writer, sheet_name=original_sheet_name, index=False)
    writer.close()
    
    return output.getvalue(), output_filename


def generate_missing_scan(raw_file_bytes, spv_file_bytes):
    df_spv_static = clean_spv_df(spv_file_bytes)

    valid_spvs = [str(spv).strip().upper() for spv in df_spv_static['SPV'].dropna().unique() if 'CEDIS' not in str(spv).strip().upper()]

    spv_map = df_spv_static[['SPV', 'ZONA', 'PDV']].dropna(subset=['PDV']).copy()
    spv_map.columns = ['spv', 'zona', 'pdv']
    spv_map['spv'] = spv_map['spv'].astype(str).str.strip()
    spv_map['zona'] = spv_map['zona'].astype(str).str.strip().replace('nan', 'N/A')
    spv_map['pdv'] = spv_map['pdv'].astype(str).str.strip()
    spv_map.drop_duplicates(subset=['pdv'], inplace=True)

    try:
        df_raw = pd.read_excel(io.BytesIO(raw_file_bytes), sheet_name='sheet0', engine='openpyxl')
    except ValueError:
        df_raw = pd.read_excel(io.BytesIO(raw_file_bytes), sheet_name=0, engine='openpyxl')
    df_raw.columns = df_raw.columns.str.strip()

    raw_date = df_raw['Fecha de estadísticas'].dropna().iloc[0]
    date_obj = pd.to_datetime(raw_date)
    months_en = {1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR', 5: 'MAY', 6: 'JUN', 7: 'JUL', 8: 'AUG', 9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DEC'}
    months_es = {1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'}
    
    # SEPARATED TITLE & SUBTITLE STRINGS
    main_title_str = "缺失扫描报告 OMISIÓN DE ESCANEO"
    subtitle_date_str = f"揽收日期 {date_obj.month}月{date_obj.day}日 Fecha de Recoleccion: {months_es[date_obj.month]} {date_obj.day}"
    output_filename = f"MISSING SCAN {months_en[date_obj.month]} {date_obj.day}.xlsx"

    val_cols = ['El número de pedidos de escaneo', 'No. de escaneo faltante de recolección', 'Nº de guías con escaneo faltantes de salida']
    pivot_raw = pd.pivot_table(df_raw, index=['Nombre del nodo'], values=val_cols, aggfunc='sum').reset_index()
    pivot_raw.rename(columns={'Nombre del nodo': 'pdv'}, inplace=True)
    pivot_raw['pdv'] = pivot_raw['pdv'].astype(str).str.strip()

    final_df = pd.merge(spv_map, pivot_raw, on='pdv', how='inner')
    for col in val_cols:
        final_df[col] = final_df[col].fillna(0)

    final_df['spv_upper'] = final_df['spv'].astype(str).str.strip().str.upper()
    final_df = final_df[final_df['spv_upper'].isin(valid_spvs)].copy()

    final_df['tasa_recoleccion'] = final_df['No. de escaneo faltante de recolección'] / final_df['El número de pedidos de escaneo']
    final_df['tasa_salida'] = final_df['Nº de guías con escaneo faltantes de salida'] / final_df['El número de pedidos de escaneo']
    final_df.fillna({'tasa_recoleccion': 0, 'tasa_salida': 0}, inplace=True)
    final_df.sort_values(by=['spv', 'zona', 'pdv'], ascending=[True, True, True], inplace=True)

    wb = openpyxl.load_workbook(io.BytesIO(raw_file_bytes))
    if 'Report' in wb.sheetnames:
        del wb['Report']
    ws = wb.create_sheet(title='Report', index=0)

    # EXACT TYPOGRAPHY SIZE REQUESTS: 36pt Title, 24pt Bold Subtitle, White Bold Totals
    font_title = Font(name='Calibri', size=36, bold=True, color='FFFFFF')
    font_subtitle_24 = Font(name='Calibri', size=24, bold=True, color='FFFFFF')
    font_white_bold = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    font_black_bold = Font(name='Calibri', size=14, bold=True, color='000000')
    font_regular = Font(name='Calibri', size=14, bold=False, color='000000')

    fill_black = PatternFill(start_color='1F1F1F', end_color='1F1F1F', fill_type='solid')
    fill_dark_grey = PatternFill(start_color='4F4F4F', end_color='4F4F4F', fill_type='solid')
    fill_mid_grey = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    fill_red = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'), top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))

    # EXPANDED COLUMN WIDTHS TO ELIMINATE CROPPING
    col_widths = {'A': 22, 'B': 22, 'C': 38, 'D': 22, 'E': 36, 'F': 36, 'G': 30, 'H': 36}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ROW 1: 36PT TITLE
    ws.merge_cells('A1:H1')
    cell_title = ws['A1']
    cell_title.value = main_title_str
    cell_title.font = font_title
    cell_title.fill = fill_black
    cell_title.alignment = align_center
    ws.row_dimensions[1].height = 50

    # ROW 2: 24PT BOLD DATE SUBTITLE
    ws.merge_cells('A2:H2')
    cell_sub = ws['A2']
    cell_sub.value = subtitle_date_str
    cell_sub.font = font_subtitle_24
    cell_sub.fill = fill_black
    cell_sub.alignment = align_center
    ws.row_dimensions[2].height = 38

    # ROW 3: TABLE HEADERS (EXACT WORDING FROM REFERENCE IMAGE)
    headers = [
        '负责人\nSPV', 
        '区域\nZONA', 
        '揽收网点\nPDV', 
        '总需扫描数量\nTOTAL A ESCANEAR', 
        '未进行入库扫描的包裹\nSIN ESCANEO DE RECOLECCION', 
        '未进行入库扫描包裹的百分比\n% SIN ESCANEO DE RECOLECCION', 
        '未进行出库扫描的包裹\nSIN ESCANEO DE SALIDA', 
        '未进行出库扫描包裹的百分比\n% SIN ESCANEO DE SALIDA'
    ]
    for col_num, h_text in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_num, value=h_text)
        c.font = font_white_bold
        c.fill = fill_dark_grey
        c.alignment = align_center
        c.border = thin_border
    ws.row_dimensions[3].height = 75

    # ROW 4: GRAND TOTAL
    gt_total = final_df['El número de pedidos de escaneo'].sum()
    gt_rec = final_df['No. de escaneo faltante de recolección'].sum()
    gt_rec_pct = gt_rec / gt_total if gt_total else 0
    gt_sal = final_df['Nº de guías con escaneo faltantes de salida'].sum()
    gt_sal_pct = gt_sal / gt_total if gt_total else 0

    ws.merge_cells('A4:C4')
    ws['A4'].value, ws['A4'].font, ws['A4'].fill, ws['A4'].alignment, ws['A4'].border = '总计', font_white_bold, fill_mid_grey, align_center, thin_border
    
    for col_idx, val, num_fmt, f_style, fill_style in [(4, gt_total, '#,##0', font_white_bold, fill_mid_grey), (5, gt_rec, '#,##0', font_white_bold, fill_mid_grey), (6, gt_rec_pct, '0.00%', font_white_bold, fill_red), (7, gt_sal, '#,##0', font_white_bold, fill_mid_grey), (8, gt_sal_pct, '0.00%', font_white_bold, fill_red)]:
        c = ws.cell(row=4, column=col_idx, value=val)
        c.font, c.fill, c.alignment, c.number_format, c.border = f_style, fill_style, align_center, num_fmt, thin_border
    ws.row_dimensions[4].height = 32

    current_row = 5
    for spv_name, group in final_df.groupby('spv', sort=False):
        spv_start_row = current_row
        
        zona_start_row = current_row
        current_zona = None

        for _, row in group.iterrows():
            ws.cell(row=current_row, column=2, value=row['zona']).alignment = align_center
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).font = font_regular
            
            if current_zona != row['zona']:
                if current_zona is not None and current_row - 1 > zona_start_row:
                    ws.merge_cells(start_row=zona_start_row, start_column=2, end_row=current_row - 1, end_column=2)
                current_zona = row['zona']
                zona_start_row = current_row

            ws.cell(row=current_row, column=3, value=row['pdv']).alignment = align_center
            for c_idx, key, fmt in [(4, 'El número de pedidos de escaneo', '#,##0'), (5, 'No. de escaneo faltante de recolección', '#,##0'), (6, 'tasa_recoleccion', '0.00%'), (7, 'Nº de guías con escaneo faltantes de salida', '#,##0'), (8, 'tasa_salida', '0.00%')]:
                c = ws.cell(row=current_row, column=c_idx, value=float(row[key]) if '%' in fmt else row[key])
                c.number_format, c.alignment = fmt, align_center
            for col_idx in [3, 4, 5, 6, 7, 8]:
                ws.cell(row=current_row, column=col_idx).font = font_regular
                ws.cell(row=current_row, column=col_idx).border = thin_border
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        if current_zona is not None and current_row - 1 > zona_start_row:
            ws.merge_cells(start_row=zona_start_row, start_column=2, end_row=current_row - 1, end_column=2)

        if current_row - 1 > spv_start_row: 
            ws.merge_cells(start_row=spv_start_row, start_column=1, end_row=current_row - 1, end_column=1)
            
        ws.cell(row=spv_start_row, column=1, value=spv_name).font = font_black_bold
        ws.cell(row=spv_start_row, column=1).alignment = align_center
        ws.cell(row=spv_start_row, column=1).border = thin_border

        sub_total = group['El número de pedidos de escaneo'].sum()
        sub_rec = group['No. de escaneo faltante de recolección'].sum()
        sub_rec_pct = sub_rec / sub_total if sub_total else 0
        sub_sal = group['Nº de guías con escaneo faltantes de salida'].sum()
        sub_sal_pct = sub_sal / sub_total if sub_total else 0

        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        sub = ws.cell(row=current_row, column=1, value=f'TOTAL {spv_name}')
        sub.font, sub.fill, sub.alignment, sub.border = font_black_bold, fill_yellow, align_center, thin_border
        
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).border = thin_border

        for col_idx, val, num_fmt, f_style, fill_style in [(4, sub_total, '#,##0', font_black_bold, fill_yellow), (5, sub_rec, '#,##0', font_black_bold, fill_yellow), (6, sub_rec_pct, '0.00%', font_white_bold, fill_red), (7, sub_sal, '#,##0', font_black_bold, fill_yellow), (8, sub_sal_pct, '0.00%', font_white_bold, fill_red)]:
            c = ws.cell(row=current_row, column=col_idx, value=val)
            c.font, c.fill, c.alignment, c.number_format, c.border = f_style, fill_style, align_center, num_fmt, thin_border
        ws.row_dimensions[current_row].height = 32
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), output_filename


def generate_r7_cdmx(raw_file_bytes, spv_file_bytes):
    df_spv = clean_spv_df(spv_file_bytes)
    
    df_spv['SPV'] = df_spv['SPV'].astype(str).str.strip().str.upper()
    df_spv['ZONA'] = df_spv['ZONA'].astype(str).str.strip()
    df_spv['PDV'] = df_spv['PDV'].astype(str).str.strip()
    df_spv['PDV_lower'] = df_spv['PDV'].str.lower()
    
    valid_spvs = [str(spv).strip().upper() for spv in df_spv['SPV'].dropna().unique() if 'CEDIS' not in str(spv).strip().upper()]
    spv_map = df_spv[df_spv['SPV'].isin(valid_spvs)].drop_duplicates(subset=['PDV_lower']).copy()

    df_raw = None
    try:
        df_raw = pd.read_excel(io.BytesIO(raw_file_bytes), engine='openpyxl')
    except Exception:
        for enc in ['utf-8-sig', 'latin1', 'cp1252', 'gbk']:
            for sep in [',', '\t', ';', '|']:
                try:
                    df_temp = pd.read_csv(io.BytesIO(raw_file_bytes), sep=sep, encoding=enc, on_bad_lines='skip', low_memory=False)
                    if len(df_temp.columns) > 3:
                        df_raw = df_temp
                        break
                except Exception: continue
            if df_raw is not None: break

    df_raw.columns = df_raw.columns.astype(str).str.strip()
    df_raw['NORMAL_PDV'] = df_raw['Punto de Recogida'].astype(str).str.strip()
    df_raw['PDV_lower'] = df_raw['NORMAL_PDV'].str.lower()
    df_raw['Estado_lower'] = df_raw['Estado de la orden'].astype(str).str.strip().str.lower()

    order_times = pd.to_datetime(df_raw['Tiempo de registro de la orden'], errors='coerce')
    unique_dates = sorted(order_times.dt.date.dropna().unique())
    months_en = {1: 'JAN', 2: 'FEB', 3: 'MAR', 4: 'APR', 5: 'MAY', 6: 'JUN', 7: 'JUL', 8: 'AUG', 9: 'SEP', 10: 'OCT', 11: 'NOV', 12: 'DEC'}
    months_es = {1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'}

    if unique_dates:
        min_o, max_o = unique_dates[0], unique_dates[-1]
        reco_date = max_o + datetime.timedelta(days=1)
        if min_o == max_o:
            pedidos_es = f"{months_es[min_o.month]} {min_o.day}"
            pedidos_zh = f"{min_o.month}月{min_o.day}日"
        elif min_o.month == max_o.month:
            pedidos_es = f"{months_es[min_o.month]} {min_o.day}-{max_o.day}"
            pedidos_zh = f"{min_o.month}月{min_o.day}-{max_o.day}日"
        else:
            pedidos_es = f"{months_es[min_o.month]} {min_o.day}-{months_es[max_o.month]} {max_o.day}"
            pedidos_zh = f"{min_o.month}月{min_o.day}日-{max_o.month}月{max_o.day}日"
        reco_es = f"{months_es[reco_date.month]} {reco_date.day}"
        reco_zh = f"{reco_date.month}月{reco_date.day}日"
    else:
        today = datetime.date.today()
        reco_date = today + datetime.timedelta(days=1)
        pedidos_es, pedidos_zh = f"{months_es[today.month]} {today.day}", f"{today.month}月{today.day}日"
        reco_es, reco_zh = f"{months_es[reco_date.month]} {reco_date.day}", f"{reco_date.month}月{reco_date.day}日"

    subtitle_str = f"订单日期: {pedidos_zh}, 揽收日期 {reco_zh} Pedidos: {pedidos_es} | Reco: {reco_es}"
    
    output_filename = f"R7 CDMX {months_en[reco_date.month]} {reco_date.day}.xlsx"

    df_total = df_raw[~df_raw['Estado_lower'].str.contains('cancelad', na=False)].copy()
    pivot_total = df_total.groupby('PDV_lower').size().reset_index(name='Total de Guias')

    spv_lookup = spv_map.set_index('PDV_lower')
    df_report = pivot_total.copy()
    df_report['SPV'] = df_report['PDV_lower'].map(spv_lookup['SPV'])
    df_report['ZONA'] = df_report['PDV_lower'].map(spv_lookup['ZONA'])
    df_report['PDV'] = df_report['PDV_lower'].map(spv_lookup['PDV'])
    df_report = df_report[df_report['SPV'].notna()].copy()

    mask_cancelado = df_raw['Estado_lower'].str.contains('cancelad', na=False)
    mask_recepcion = df_raw['Estado_lower'].str.contains('recep', na=False)
    df_por_rec = df_raw[~(mask_cancelado | mask_recepcion)].copy()
    pivot_por_rec = df_por_rec.groupby('PDV_lower').size().reset_index(name='Guias por Recolectar')

    por_rec_lookup = pivot_por_rec.set_index('PDV_lower')['Guias por Recolectar']
    df_report['Guias por Recolectar'] = df_report['PDV_lower'].map(por_rec_lookup).fillna(0)
    df_report['Guias Recolectadas'] = df_report['Total de Guias'] - df_report['Guias por Recolectar']
    df_report['Rate %'] = (df_report['Guias Recolectadas'] / df_report['Total de Guias']).fillna(0)
    df_report.sort_values(by=['SPV', 'ZONA', 'PDV'], ascending=[True, True, True], inplace=True)

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    wb = writer.book
    ws = wb.add_worksheet('R7 CDMX')

    dark_purple, light_purple, yellow, red = '#2E003E', '#480060', '#FFFF00', '#990000'
    base_font = 'Calibri'
    base_size = 14
    
    t_fmt = wb.add_format({'bold': True, 'font_size': 36, 'align': 'center', 'valign': 'vcenter', 'bg_color': dark_purple, 'font_color': 'white', 'font_name': base_font})
    # BOLD 24PT DATE SUBTITLE
    s_fmt = wb.add_format({'bold': True, 'font_size': 24, 'align': 'center', 'valign': 'vcenter', 'bg_color': dark_purple, 'font_color': 'white', 'font_name': base_font})
    h_fmt = wb.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': dark_purple, 'font_color': 'white', 'border': 1, 'text_wrap': True, 'font_name': base_font})
    
    gt_l = wb.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': light_purple, 'font_color': 'white', 'border': 1, 'font_name': base_font})
    gt_v = wb.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': light_purple, 'font_color': 'white', 'border': 1, 'num_format': '#,##0', 'font_name': base_font})
    gt_p = wb.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': red, 'font_color': 'white', 'border': 1, 'num_format': '0.00%', 'font_name': base_font})
    
    sub_l = wb.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': yellow, 'font_color': 'black', 'border': 1, 'font_name': base_font})
    sub_v = wb.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': yellow, 'font_color': 'black', 'border': 1, 'num_format': '#,##0', 'font_name': base_font})
    sub_p = wb.add_format({'bold': True, 'font_size': base_size, 'align': 'center', 'valign': 'vcenter', 'bg_color': red, 'font_color': 'white', 'border': 1, 'num_format': '0.00%', 'font_name': base_font})
    
    spv_f = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'bold': True, 'font_size': base_size, 'font_name': base_font})
    d_f = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': base_size, 'num_format': '#,##0', 'font_name': base_font})
    dp_f = wb.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': base_size, 'num_format': '0.00%', 'font_name': base_font})

    ws.set_column('A:A', 22); ws.set_column('B:B', 22); ws.set_column('C:C', 36); ws.set_column('D:F', 26); ws.set_column('G:G', 24)
    ws.merge_range('A1:G1', 'R7 CDMX 所有平台的揽收率', t_fmt)
    ws.set_row(0, 80) 
    ws.merge_range('A2:G2', subtitle_str, s_fmt)
    ws.set_row(1, 38)

    headers = ['', 'ZONA', '客户归属网点\nPDV', '当日包裹总数\nTotal de Guias', '已收取的包裹\nGuias Recolectadas', '待收取的包裹\nGuias por Recolectar', '商家拜访率\nRate %']
    for col, h in enumerate(headers): ws.write(2, col, h, h_fmt)
    ws.set_row(2, 60)

    ws.merge_range('A4:C4', 'GRAND TOTAL 总计', gt_l)
    ws.write(3, 3, df_report['Total de Guias'].sum(), gt_v)
    ws.write(3, 4, df_report['Guias Recolectadas'].sum(), gt_v)
    ws.write(3, 5, df_report['Guias por Recolectar'].sum(), gt_v)
    ws.write(3, 6, df_report['Guias Recolectadas'].sum() / df_report['Total de Guias'].sum() if df_report['Total de Guias'].sum() else 0, gt_p)
    ws.set_row(3, 34)

    c_row = 4
    for spv_name, group in df_report.groupby('SPV', sort=False):
        spv_start_row = c_row
        
        zona_start_row = c_row
        current_zona = None

        for _, row in group.iterrows():
            if current_zona != row['ZONA']:
                if current_zona is not None:
                    if c_row - 1 > zona_start_row:
                        ws.merge_range(zona_start_row, 1, c_row - 1, 1, current_zona, d_f)
                    else:
                        ws.write(zona_start_row, 1, current_zona, d_f)
                current_zona = row['ZONA']
                zona_start_row = c_row

            ws.write(c_row, 2, row['PDV'], d_f)
            ws.write(c_row, 3, row['Total de Guias'], d_f)
            ws.write(c_row, 4, row['Guias Recolectadas'], d_f)
            
            # BLANK WHEN ZERO FOR VISUAL CLEANLINESS
            por_rec_val = row['Guias por Recolectar']
            ws.write(c_row, 5, "" if por_rec_val == 0 else por_rec_val, d_f)
            
            ws.write(c_row, 6, float(row['Rate %']), dp_f)
            ws.set_row(c_row, 24)
            c_row += 1
            
        if current_zona is not None:
            if c_row - 1 > zona_start_row:
                ws.merge_range(zona_start_row, 1, c_row - 1, 1, current_zona, d_f)
            else:
                ws.write(zona_start_row, 1, current_zona, d_f)

        if c_row - 1 > spv_start_row: 
            ws.merge_range(spv_start_row, 0, c_row - 1, 0, spv_name, spv_f)
        else: 
            ws.write(spv_start_row, 0, spv_name, spv_f)

        s_tot, s_rec, s_por = group['Total de Guias'].sum(), group['Guias Recolectadas'].sum(), group['Guias por Recolectar'].sum()
        ws.merge_range(f'A{c_row+1}:C{c_row+1}', f'TOTAL {spv_name}', sub_l)
        ws.write(c_row, 3, s_tot, sub_v)
        ws.write(c_row, 4, s_rec, sub_v)
        ws.write(c_row, 5, s_por, sub_v)
        ws.write(c_row, 6, s_rec/s_tot if s_tot else 0, sub_p)
        ws.set_row(c_row, 34)
        c_row += 1

    writer.close()
    return output.getvalue(), output_filename


def generate_anomalies(raw_file_bytes, spv_file_bytes, raw_filename):
    df_spv = clean_spv_df(spv_file_bytes)

    df_spv['SPV'] = df_spv['SPV'].astype(str).str.strip().str.upper()
    df_spv['ZONA'] = df_spv['ZONA'].astype(str).str.strip()
    df_spv['PDV'] = df_spv['PDV'].astype(str).str.strip()
    df_spv['PDV_lower'] = df_spv['PDV'].str.lower()
    
    valid_spvs = [str(spv).strip().upper() for spv in df_spv['SPV'].dropna().unique() if 'CEDIS' not in str(spv).strip().upper()]
    spv_map = df_spv[df_spv['SPV'].isin(valid_spvs)].drop_duplicates(subset=['PDV_lower']).copy()

    xls = pd.ExcelFile(io.BytesIO(raw_file_bytes), engine='openpyxl')
    target_sheet = [s for s in xls.sheet_names if '未取件客户明细' in s or '网点明细' in s]
    sheet_to_load = target_sheet[0] if target_sheet else 0
    df_raw = pd.read_excel(io.BytesIO(raw_file_bytes), sheet_name=sheet_to_load, engine='openpyxl')
    df_raw.columns = df_raw.columns.astype(str).str.strip()

    pdv_col = [c for c in df_raw.columns if '网点' in c][0]
    pendientes_col = [c for c in df_raw.columns if '未取件订单' in c][0]
    registrados_col = [c for c in df_raw.columns if '已登记' in c][0]
    no_registrados_col = [c for c in df_raw.columns if '未登记' in c][0]

    order_col = [c for c in df_raw.columns if '订单日期' in c]
    reco_col = [c for c in df_raw.columns if '考核' in c or '收件' in c]

    order_times = pd.to_datetime(df_raw[order_col[0]], errors='coerce') if order_col else pd.Series(dtype='datetime64[ns]')
    reco_times = pd.to_datetime(df_raw[reco_col[0]], errors='coerce') if reco_col else pd.Series(dtype='datetime64[ns]')
    unique_order = sorted(order_times.dt.date.dropna().unique())
    unique_reco = sorted(reco_times.dt.date.dropna().unique())
    months_es = {1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'}

    if unique_order and unique_reco:
        o_date, r_date = unique_order[-1], unique_reco[-1]
        pedidos_zh, pedidos_es = f"{o_date.month}月{o_date.day}日", f"{months_es[o_date.month]} {o_date.day}"
        reco_zh, reco_es = f"{r_date.month}月{r_date.day}日", f"{months_es[r_date.month]} {r_date.day}"
    else:
        today = datetime.date.today()
        r_date = today + datetime.timedelta(days=1)
        pedidos_zh, pedidos_es = f"{today.month}月{today.day}日", f"{months_es[today.month]} {today.day}"
        reco_zh, reco_es = f"{r_date.month}月{r_date.day}日", f"{months_es[r_date.month]} {r_date.day}"

    subtitle_str = f"订单日期： {pedidos_zh}, 收件日期 {reco_zh} Pedidos: {pedidos_es} | Reco: {reco_es}"
    
    output_filename = raw_filename if raw_filename.endswith('.xlsx') else f"{raw_filename.split('.')[0]}.xlsx"

    for col in [pendientes_col, registrados_col, no_registrados_col]:
        df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce').fillna(0)

    pivot_total = df_raw.groupby(pdv_col).agg({pendientes_col: 'sum', registrados_col: 'sum', no_registrados_col: 'sum'}).reset_index()
    pivot_total.columns = ['PDV', 'Pendientes', 'Registrados', 'No_Registrados']
    pivot_total['PDV_lower'] = pivot_total['PDV'].astype(str).str.strip().str.lower()

    spv_lookup = spv_map.set_index('PDV_lower')
    df_report = pivot_total.copy()
    df_report['SPV'] = df_report['PDV_lower'].map(spv_lookup['SPV'])
    df_report['ZONA'] = df_report['PDV_lower'].map(spv_lookup['ZONA'])
    df_report['PDV_Clean'] = df_report['PDV_lower'].map(spv_lookup['PDV'])

    df_report = df_report[df_report['SPV'].notna()].copy()
    df_report['PDV'] = df_report['PDV_Clean'].combine_first(df_report['PDV'])
    df_report['Rate %'] = (df_report['Registrados'] / df_report['Pendientes']).fillna(0)
    df_report.sort_values(by=['SPV', 'ZONA', 'No_Registrados', 'PDV'], ascending=[True, True, False, True], inplace=True)

    wb = openpyxl.load_workbook(io.BytesIO(raw_file_bytes))
    if 'Anomalias' in wb.sheetnames: del wb['Anomalias']
    ws = wb.create_sheet(title='Anomalias', index=0)

    font_white_bold = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    font_black_bold = Font(name='Calibri', size=14, bold=True, color='000000')
    font_regular = Font(name='Calibri', size=14, bold=False, color='000000')
    font_subtitle_24 = Font(name='Calibri', size=24, bold=True, color='FFFFFF')
    
    fil_main = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    fil_dark_grey = PatternFill(start_color='333333', end_color='333333', fill_type='solid')
    fil_y = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    a_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'), top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))

    c_w = {'A': 22, 'B': 22, 'C': 36, 'D': 28, 'E': 28, 'F': 28, 'G': 26}
    for col, width in c_w.items(): ws.column_dimensions[col].width = width

    ws.merge_cells('A1:G1')
    ws['A1'].value, ws['A1'].font, ws['A1'].fill, ws['A1'].alignment = '问题件跟进 Seguimiento Paquetes de Anomalia', Font(name='Calibri', size=36, bold=True, color='FFFFFF'), fil_main, a_c
    ws.row_dimensions[1].height = 50 

    # BOLD 24PT DATE SUBTITLE
    ws.merge_cells('A2:G2')
    ws['A2'].value, ws['A2'].font, ws['A2'].fill, ws['A2'].alignment = subtitle_str, font_subtitle_24, fil_main, a_c
    ws.row_dimensions[2].height = 38

    headers = ['', 'ZONA', '客户归属网点\nPDV', '未取件订单量合计\nPaquetes pendientes de Recoleccion', '已登记问题件量合计\nPaquetes de Anomalia Registrados', '未登记问题件量合计\nPaquetes de Anomalia NO Registrados', '问题件登记率\n% Registro de Paquetes de Anomalia']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = font_white_bold, fil_dark_grey, a_c, thin_border
    ws.row_dimensions[3].height = 65

    # GRAND TOTAL ROW: QUANTITIES IN DARK GREY, PERCENTAGE RATE ONLY IN RED
    gt_p, gt_r, gt_n = df_report['Pendientes'].sum(), df_report['Registrados'].sum(), df_report['No_Registrados'].sum()
    
    ws.merge_cells('A4:C4')
    ws['A4'].value, ws['A4'].font, ws['A4'].fill, ws['A4'].alignment, ws['A4'].border = 'GRAND TOTAL 总计', font_white_bold, fil_dark_grey, a_c, thin_border

    for i, val in enumerate([gt_p, gt_r, gt_n], start=4):
        c = ws.cell(row=4, column=i, value=val)
        c.font, c.fill, c.alignment, c.number_format, c.border = font_white_bold, fil_dark_grey, a_c, '#,##0', thin_border
        
    # PERCENTAGE CELL HIGHLIGHTED RED (fil_main) ONLY
    c_gt_rate = ws.cell(row=4, column=7, value=gt_r/gt_p if gt_p else 0)
    c_gt_rate.font, c_gt_rate.fill, c_gt_rate.alignment, c_gt_rate.number_format, c_gt_rate.border = font_white_bold, fil_main, a_c, '0.00%', thin_border
    ws.row_dimensions[4].height = 34

    current_row = 5
    for spv_name, group in df_report.groupby('SPV', sort=False):
        spv_start_row = current_row
        
        zona_start_row = current_row
        current_zona = None

        for _, row in group.iterrows():
            ws.cell(row=current_row, column=2, value=row['ZONA']).alignment = a_c
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).font = font_regular
            
            if current_zona != row['ZONA']:
                if current_zona is not None and current_row - 1 > zona_start_row:
                    ws.merge_cells(start_row=zona_start_row, start_column=2, end_row=current_row - 1, end_column=2)
                current_zona = row['ZONA']
                zona_start_row = current_row

            ws.cell(row=current_row, column=3, value=row['PDV']).alignment = a_c
            
            # BLANK ZERO VALUES FOR INDIVIDUAL PDV DATA ROWS
            v_pend = row['Pendientes']
            v_reg = row['Registrados']
            v_noreg = row['No_Registrados']
            
            disp_pend = "" if v_pend == 0 else v_pend
            disp_reg = "" if v_reg == 0 else v_reg
            disp_noreg = "" if v_noreg == 0 else v_noreg
            
            ws.cell(row=current_row, column=4, value=disp_pend).number_format = '#,##0'
            ws.cell(row=current_row, column=5, value=disp_reg).number_format = '#,##0'
            ws.cell(row=current_row, column=6, value=disp_noreg).number_format = '#,##0'
            
            c_rate = ws.cell(row=current_row, column=7, value=float(row['Rate %']))
            c_rate.number_format, c_rate.alignment = '0.00%', a_c

            for col_idx in [3, 4, 5, 6, 7]:
                ws.cell(row=current_row, column=col_idx).font = font_regular
                ws.cell(row=current_row, column=col_idx).border = thin_border
                ws.cell(row=current_row, column=col_idx).alignment = a_c
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        if current_zona is not None and current_row - 1 > zona_start_row:
            ws.merge_cells(start_row=zona_start_row, start_column=2, end_row=current_row - 1, end_column=2)

        if current_row - 1 > spv_start_row: 
            ws.merge_cells(start_row=spv_start_row, start_column=1, end_row=current_row - 1, end_column=1)
            
        c = ws.cell(row=spv_start_row, column=1, value=spv_name)
        c.font, c.alignment, c.border = font_black_bold, a_c, thin_border

        s_p, s_r, s_n = group['Pendientes'].sum(), group['Registrados'].sum(), group['No_Registrados'].sum()
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        c = ws.cell(row=current_row, column=1, value=f'TOTAL {spv_name}')
        c.font, c.fill, c.alignment, c.border = font_black_bold, fil_y, a_c, thin_border
        
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=3).border = thin_border

        for col, val in [(4, s_p), (5, s_r), (6, s_n)]:
            c = ws.cell(row=current_row, column=col, value=val)
            c.font, c.fill, c.alignment, c.number_format, c.border = font_black_bold, fil_y, a_c, '#,##0', thin_border

        # PERCENTAGE RATE CELL ONLY IN RED (fil_main)
        c_sub_rate = ws.cell(row=current_row, column=7, value=s_r/s_p if s_p else 0)
        c_sub_rate.font, c_sub_rate.fill, c_sub_rate.alignment, c_sub_rate.number_format, c_sub_rate.border = font_white_bold, fil_main, a_c, '0.00%', thin_border
            
        ws.row_dimensions[current_row].height = 34
        current_row += 1

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), output_filename


# ==============================================================================
# MAIN WEB UI
# ==============================================================================
if check_password():
    st.title("📊 Automated Reporting Portal")
    st.markdown("Upload your daily raw data below to instantly generate standardized Excel reports.")
    
    # 1. READ "SUPERVISORES VISITAS.xlsx" FOR TIKTOK
    try:
        with open("SUPERVISORES VISITAS.xlsx", "rb") as f:
            spv_visitas_bytes = f.read()
    except FileNotFoundError:
        st.error(
            "🚨 **CRITICAL ERROR:** `SUPERVISORES VISITAS.xlsx` was not found in the root directory. "
            "Please make sure the file is saved in the repository on GitHub."
        )
        st.stop()

    # 2. READ "SUPERVISORES.xlsx" FOR EVERYTHING ELSE
    try:
        with open("SUPERVISORES.xlsx", "rb") as f:
            spv_general_bytes = f.read()
    except FileNotFoundError:
        st.error(
            "🚨 **CRITICAL ERROR:** `SUPERVISORES.xlsx` was not found in the root directory. "
            "Please make sure the file is saved in the repository on GitHub."
        )
        st.stop()
        
    st.divider()

    report_type = st.selectbox(
        "Select the report you want to generate:",
        ("MISSING SCAN", "R7 CDMX", "ANOMALIES (问题件跟进)", "ALIEXPRESS", "TIKTOK PENDING VISITS")
    )
    
    raw_data = st.file_uploader("📥 Upload Daily Raw Data (.xlsx, .csv)", type=["xlsx", "csv"])
    
    if st.button("🚀 Generate Report", type="primary", use_container_width=True):
        if raw_data is None:
            st.warning("⚠️ Please upload the Daily Raw Data before proceeding.")
        else:
            with st.spinner(f'Crunching the numbers and formatting your {report_type} report...'):
                try:
                    if report_type == "TIKTOK PENDING VISITS":
                        processed_file, filename = generate_tiktok_visits(raw_data.getvalue(), spv_visitas_bytes)
                    elif report_type == "MISSING SCAN":
                        processed_file, filename = generate_missing_scan(raw_data.getvalue(), spv_general_bytes)
                    elif report_type == "R7 CDMX":
                        processed_file, filename = generate_r7_cdmx(raw_data.getvalue(), spv_general_bytes)
                    elif report_type == "ANOMALIES (问题件跟进)":
                        processed_file, filename = generate_anomalies(raw_data.getvalue(), spv_general_bytes, raw_data.name)
                    elif report_type == "ALIEXPRESS":
                        processed_file, filename = generate_aliexpress(raw_data.getvalue(), spv_general_bytes)
                    
                    st.success(f"✅ Your {report_type} report was generated successfully!")
                    
                    st.download_button(
                        label="📥 Download Finished Excel Report",
                        data=processed_file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"❌ An error occurred while processing the file: {e}")